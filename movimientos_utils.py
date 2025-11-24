"""Utilidades comunes para la manipulacion del Excel de movimientos."""

from __future__ import annotations

from typing import Optional, Sequence

import pandas as pd
from openpyxl import load_workbook

COLUMNAS_TEXTO_FORZADO = [
    "Nota de Venta",
    "Orden de Compra",
    "Codigo",
    "Cantidad",
]


def preparar_movimientos_para_excel(df_mov: pd.DataFrame) -> pd.DataFrame:
    """Convierte las columnas sensibles a texto antes de exportar."""
    df_salida = df_mov.copy()
    for columna in COLUMNAS_TEXTO_FORZADO:
        if columna in df_salida.columns:
            df_salida[columna] = df_salida[columna].fillna("").astype(str)
    return df_salida


def guardar_movimientos_excel(df_mov: pd.DataFrame, ruta: str):
    """Guarda el DataFrame asegurando formato texto en las columnas definidas."""
    df_salida = preparar_movimientos_para_excel(df_mov)
    df_salida.to_excel(ruta, index=False)
    _forzar_formato_texto(ruta, df_salida.columns)


def forzar_columnas_texto_excel(ruta: str, columnas_presentes: Sequence[str], columnas_objetivo: Sequence[str]):
    """Expone el formateo de columnas a texto para otros Excel fuera de movimientos."""
    _forzar_formato_texto(ruta, columnas_presentes, columnas_objetivo)


def _forzar_formato_texto(
    ruta: str, columnas: Sequence[str], columnas_objetivo: Optional[Sequence[str]] = None
):
    objetivos = columnas_objetivo if columnas_objetivo is not None else COLUMNAS_TEXTO_FORZADO
    libro = load_workbook(ruta)
    hoja = libro.active
    for columna in objetivos:
        if columna not in columnas:
            continue
        idx = columnas.index(columna) + 1 if isinstance(columnas, list) else columnas.get_loc(columna) + 1
        for fila in range(2, hoja.max_row + 1):
            celda = hoja.cell(row=fila, column=idx)
            celda.number_format = "@"
    libro.save(ruta)
